from io import TextIOWrapper
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv

def test_zip_files(download_files, zip_files):
    with ZipFile("resources/ZZZIPED.zip") as zip_file:
        assert len(zip_file.filelist) == 3
        assert (zip_file.namelist()) == ['tmp/XLSX_test.xlsx', 'tmp/TestPdf.pdf', 'tmp/CSV_test.CSV']

def test_pdf_file():
    with ZipFile("resources/ZZZIPED.zip") as zip_file:
        with zip_file.open("tmp/TestPdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            test_text = reader.pages[0].extract_text()
            number_pages = (len(reader.pages))
            print(test_text)
            print(number_pages)
            assert "С уважением, ООО «СК «Согласие»" in test_text
            assert number_pages == 3

def test_xlsx_file():
    with ZipFile("resources/ZZZIPED.zip") as zip_file:
        with zip_file.open("tmp/XLSX_test.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            check_cell = sheet.cell(row=1, column=2).value
            assert check_cell == "Вышестоящий отдел"

def test_csv_file():
    with ZipFile("resources/ZZZIPED.zip") as zip_file:
        with zip_file.open("tmp/CSV_test.CSV") as csv_file:
            reader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8')))
            assert ['1', 'Первый', 'One'] == reader[1]